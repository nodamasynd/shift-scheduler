from flask import Flask, render_template, request, jsonify, send_file
from ortools.sat.python import cp_model
import pandas as pd
from datetime import datetime, timedelta
import io
import json

app = Flask(__name__)

class ShiftScheduler:
    def __init__(self, month, year, num_staff, staff_info, requests_off, preferred_shifts, holidays, early_count, middle_count, late_count_min, late_count_max, balance_tolerance=2):
        self.month = month
        self.year = year
        self.num_staff = num_staff
        self.staff_info = staff_info  # {staff_id: {'name': 'xx', 'is_newbie': True/False, 'nakaban_only': True/False}}
        self.requests_off = requests_off  # {staff_id: [day1, day2, ...]}
        self.preferred_shifts = preferred_shifts  # {staff_id: {day: shift_type}}
        self.holidays = holidays  # {staff_id: [paid_holiday_days]}
        self.early_count = early_count  # 早番の必要人数
        self.middle_count = middle_count  # 中番の必要人数
        self.late_count_min = late_count_min  # 遅番の最小人数
        self.late_count_max = late_count_max  # 遅番の最大人数
        self.balance_tolerance = balance_tolerance  # 早番・遅番バランスの許容差
        
        # 月の日数を計算
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        last_day = next_month - timedelta(days=1)
        self.num_days = last_day.day
        
        # シフトタイプ: 0=早番, 1=中番, 2=遅番, 3=休み
        self.shifts = ['早番', '中番', '遅番', 'OFF']
        self.shift_times = {
            '早番': '6:25～15:25',
            '中番': '8:00～17:00',
            '遅番': '14:00～23:00'
        }
        self.EARLY = 0
        self.MIDDLE = 1
        self.LATE = 2
        self.OFF = 3
        
    def create_schedule(self, relax_balance=0, relax_preferences=False, relax_consecutive=False, relax_late_early=False):
        """
        シフトを作成する。制約を段階的に緩和して解を探す。
        
        Args:
            relax_balance: バランス制約の緩和度 (0=緩和なし, 1=+1日, 2=+2日, ...)
            relax_preferences: 希望シフトをソフト制約化するか
            relax_consecutive: 連続勤務制約を4日まで緩和するか
            relax_late_early: 遅番→早番制約を緩和するか
        """
        model = cp_model.CpModel()
        
        # 診断情報を保存
        self.diagnostics = {
            'total_staff': self.num_staff,
            'required_daily': self.early_count + self.middle_count + self.late_count_min,
            'required_daily_max': self.early_count + self.middle_count + self.late_count_max,
            'target_off_days': 8 if self.month % 2 == 0 else 9,
            'newbies': sum(1 for s in range(self.num_staff) if self.staff_info[s].get('is_newbie', False)),
            'nakaban_only': sum(1 for s in range(self.num_staff) if self.staff_info[s].get('nakaban_only', False)),
            'total_requests': sum(len(self.requests_off.get(s, [])) + len(self.preferred_shifts.get(s, {})) for s in range(self.num_staff))
        }
        
        # 変数の作成: shifts[(s, d, shift)] = スタッフsが日dにshiftで勤務するか
        shifts = {}
        for s in range(self.num_staff):
            for d in range(self.num_days):
                for shift in range(4):  # 0:早番, 1:中番, 2:遅番, 3:休み
                    shifts[(s, d, shift)] = model.NewBoolVar(f'shift_s{s}_d{d}_t{shift}')
        
        # 制約1: 各スタッフは1日に1つのシフトのみ
        for s in range(self.num_staff):
            for d in range(self.num_days):
                model.Add(sum(shifts[(s, d, shift)] for shift in range(4)) == 1)
        
        # 制約2: 各シフトの人数（Webから指定）
        late_is_three_vars = []  # 遅番が3名の日を記録
        late_is_two_vars = []    # 遅番が2名の日を記録
        late_is_four_plus_vars = []  # 遅番が4名以上の日を記録
        
        for d in range(self.num_days):
            # 早番
            model.Add(sum(shifts[(s, d, self.EARLY)] for s in range(self.num_staff)) == self.early_count)
            # 中番
            model.Add(sum(shifts[(s, d, self.MIDDLE)] for s in range(self.num_staff)) == self.middle_count)
            # 遅番（最低2名、上限なし、できるだけ3名を優先）
            late_count = sum(shifts[(s, d, self.LATE)] for s in range(self.num_staff))
            model.Add(late_count >= self.late_count_min)
            # 上限は設定しない（柔軟性を持たせる）
            if self.late_count_max > 0:
                model.Add(late_count <= self.late_count_max)
            
            # 遅番が3名の日を示すブール変数
            is_three = model.NewBoolVar(f'late_is_three_d{d}')
            model.Add(late_count == 3).OnlyEnforceIf(is_three)
            model.Add(late_count != 3).OnlyEnforceIf(is_three.Not())
            late_is_three_vars.append(is_three)
            
            # 遅番が2名の日を示すブール変数（ペナルティ）
            is_two = model.NewBoolVar(f'late_is_two_d{d}')
            model.Add(late_count == 2).OnlyEnforceIf(is_two)
            model.Add(late_count != 2).OnlyEnforceIf(is_two.Not())
            late_is_two_vars.append(is_two)
            
            # 遅番が4名以上の日を示すブール変数（ペナルティ）
            is_four_plus = model.NewBoolVar(f'late_is_four_plus_d{d}')
            model.Add(late_count >= 4).OnlyEnforceIf(is_four_plus)
            model.Add(late_count < 4).OnlyEnforceIf(is_four_plus.Not())
            late_is_four_plus_vars.append(is_four_plus)
        
        # 制約3: 中番専任スタッフは中番または休みのみ
        for s in range(self.num_staff):
            if self.staff_info[s].get('nakaban_only', False):
                for d in range(self.num_days):
                    model.Add(shifts[(s, d, self.EARLY)] == 0)
                    model.Add(shifts[(s, d, self.LATE)] == 0)
        
        # 制約4: 連続勤務は3日まで（relax_consecutive=True の場合4日まで）
        max_consecutive = 4 if relax_consecutive else 3
        for s in range(self.num_staff):
            for d in range(self.num_days - max_consecutive):
                model.Add(sum(shifts[(s, d + i, self.OFF)] for i in range(max_consecutive + 1)) >= 1)
        
        # 制約5: 遅番の次の日は早番にならない（relax_late_early=True の場合は緩和）
        if not relax_late_early:
            for s in range(self.num_staff):
                for d in range(self.num_days - 1):
                    model.Add(shifts[(s, d, self.LATE)] + shifts[(s, d + 1, self.EARLY)] <= 1)
        
        # 制約6: 新人同士は同一シフト勤務しない
        newbies = [s for s in range(self.num_staff) if self.staff_info[s].get('is_newbie', False)]
        if len(newbies) >= 2:
            for d in range(self.num_days):
                for shift in [self.EARLY, self.MIDDLE, self.LATE]:
                    model.Add(sum(shifts[(s, d, shift)] for s in newbies) <= 1)
        
        # 制約7: 希望休・希望シフト
        for s in range(self.num_staff):
            # 希望休
            if s in self.requests_off:
                for day in self.requests_off[s]:
                    if 1 <= day <= self.num_days:
                        model.Add(shifts[(s, day - 1, self.OFF)] == 1)
            
            # 有給
            if s in self.holidays:
                for day in self.holidays[s]:
                    if 1 <= day <= self.num_days:
                        model.Add(shifts[(s, day - 1, self.OFF)] == 1)
            
            # 希望シフト
            if s in self.preferred_shifts and not relax_preferences:
                # ハード制約として希望シフトを強制
                for day, shift_type in self.preferred_shifts[s].items():
                    if 1 <= day <= self.num_days:
                        # 早番、中番、遅番のみ対応
                        if shift_type in ['早番', '中番', '遅番']:
                            shift_idx = ['早番', '中番', '遅番'].index(shift_type)
                            model.Add(shifts[(s, day - 1, shift_idx)] == 1)
        
        # 制約8: 休日数（偶数月8日、奇数月9日、固定）
        target_off_days = 8 if self.month % 2 == 0 else 9
        
        for s in range(self.num_staff):
            off_count = sum(shifts[(s, d, self.OFF)] for d in range(self.num_days))
            model.Add(off_count == target_off_days)  # 固定
        
        # 制約9: 早番と遅番のバランス（中番専任以外）
        # 各スタッフの早番と遅番の回数がほぼ同じになるように（ハード制約）
        adjusted_balance_tolerance = self.balance_tolerance + relax_balance
        for s in range(self.num_staff):
            # 中番専任スタッフはスキップ
            if not self.staff_info[s].get('nakaban_only', False):
                # 早番の回数
                early_count = sum(shifts[(s, d, self.EARLY)] for d in range(self.num_days))
                # 遅番の回数
                late_count = sum(shifts[(s, d, self.LATE)] for d in range(self.num_days))
                
                # 早番と遅番の差を±(balance_tolerance + relax_balance)日以内に制限（ハード制約）
                model.Add(early_count - late_count <= adjusted_balance_tolerance)
                model.Add(late_count - early_count <= adjusted_balance_tolerance)
        
        # 目的関数: 希望シフトを最大限考慮 + 遅番3名を優先
        objective_prefs = []
        # relax_preferences=True の場合、希望シフトをソフト制約として目的関数に追加
        for s in range(self.num_staff):
            if s in self.preferred_shifts:
                for day, shift_type in self.preferred_shifts[s].items():
                    if 1 <= day <= self.num_days:
                        if shift_type in ['早番', '中番', '遅番']:
                            shift_idx = ['早番', '中番', '遅番'].index(shift_type)
                            objective_prefs.append(shifts[(s, day - 1, shift_idx)])
        
        # 目的関数: 希望達成（重み10）+ 遅番3名の日を増やす（重み5）- 遅番2名の日（重み-3）- 遅番4名以上の日（重み-3）
        objective_terms = []
        if objective_prefs:
            objective_terms.extend([10 * pref for pref in objective_prefs])
        # 遅番が3名の日を優先（希望より低い優先度）
        objective_terms.extend([5 * is_three for is_three in late_is_three_vars])
        # 遅番が2名の日を抑制（ペナルティ）
        objective_terms.extend([-3 * is_two for is_two in late_is_two_vars])
        # 遅番が4名以上の日を抑制（ペナルティ）
        objective_terms.extend([-3 * is_four_plus for is_four_plus in late_is_four_plus_vars])
        
        if objective_terms:
            model.Maximize(sum(objective_terms))
        
        # ソルバー実行
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 30.0
        status = solver.Solve(model)
        
        if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
            # 結果を取得
            schedule = {}
            off_counters = {}  # 各スタッフの休みカウンター
            
            for s in range(self.num_staff):
                schedule[s] = []
                off_counters[s] = 1  # OFF1から始める
                
                for d in range(self.num_days):
                    for shift in range(4):
                        if solver.Value(shifts[(s, d, shift)]) == 1:
                            if shift == self.OFF:
                                # 休みに番号を付ける
                                schedule[s].append(f'OFF{off_counters[s]}')
                                off_counters[s] += 1
                            else:
                                schedule[s].append(self.shifts[shift])
                            break
            return schedule, True, None
        else:
            # 失敗理由を生成
            reasons = self._analyze_failure()
            return None, False, reasons
    
    def create_schedule_with_relaxation(self):
        """
        制約を段階的に緩和してシフト作成を試みる
        
        Returns:
            (schedule, success, reasons, relaxation_info)
        """
        relaxation_attempts = [
            # (relax_balance, relax_preferences, relax_consecutive, relax_late_early, 説明)
            (0, False, False, False, '通常モード（制約緩和なし）'),
            (1, False, False, False, 'バランス+1日緩和'),
            (2, False, False, False, 'バランス+2日緩和'),
            (0, True, False, False, '希望シフトをソフト制約化'),
            (1, True, False, False, 'バランス+1日 + 希望シフトソフト化'),
            (2, True, False, False, 'バランス+2日 + 希望シフトソフト化'),
            (0, False, True, False, '連続勤務を4日まで緩和'),
            (1, False, True, False, 'バランス+1日 + 連続勤務緩和'),
            (2, True, True, False, 'バランス+2日 + 希望ソフト + 連続緩和'),
            (0, False, False, True, '遅番→早番制約を緩和'),
            (1, True, False, True, 'バランス+1日 + 希望ソフト + 遅番→早番緩和'),
            (2, True, True, False, 'バランス+2日 + 希望ソフト + 連続緩和'),
            (3, True, True, False, 'バランス+3日 + 希望ソフト + 連続緩和'),
            (2, True, True, True, 'バランス+2日 + 希望ソフト + 連続緩和 + 遅番→早番緩和'),
            (3, True, True, True, 'バランス+3日 + 希望ソフト + 連続緩和 + 遅番→早番緩和'),
            (4, True, True, True, '最大緩和（バランス+4日 + すべての制約緩和）'),
        ]
        
        for relax_balance, relax_preferences, relax_consecutive, relax_late_early, description in relaxation_attempts:
            schedule, success, reasons = self.create_schedule(
                relax_balance=relax_balance,
                relax_preferences=relax_preferences,
                relax_consecutive=relax_consecutive,
                relax_late_early=relax_late_early
            )
            
            if success:
                # 成功した場合、緩和情報を返す
                relaxation_info = None
                if relax_balance > 0 or relax_preferences or relax_consecutive or relax_late_early:
                    relaxation_info = {
                        'applied': True,
                        'description': description,
                        'relax_balance': relax_balance,
                        'relax_preferences': relax_preferences,
                        'relax_consecutive': relax_consecutive,
                        'relax_late_early': relax_late_early,
                        'adjusted_balance_tolerance': self.balance_tolerance + relax_balance
                    }
                return schedule, True, None, relaxation_info
        
        # すべて失敗した場合
        reasons = self._analyze_failure()
        return None, False, reasons, None
    
    def _analyze_failure(self):
        """シフト作成失敗の理由を詳細に分析し、具体的な修正アドバイスを提供"""
        reasons = []
        
        # 基本情報の計算
        min_required = self.early_count + self.middle_count + self.late_count_min
        max_required = self.early_count + self.middle_count + self.late_count_max
        target_off_days = 8 if self.month % 2 == 0 else 9
        working_days = self.num_days - target_off_days
        
        # カウント情報
        nakaban_only_count = sum(1 for s in range(self.num_staff) if self.staff_info[s].get('nakaban_only', False))
        newbie_count = sum(1 for s in range(self.num_staff) if self.staff_info[s].get('is_newbie', False))
        non_nakaban_staff = self.num_staff - nakaban_only_count
        
        # 1. スタッフ数不足の詳細チェック
        if self.num_staff < min_required:
            shortage = min_required - self.num_staff
            reasons.append({
                'type': 'critical',
                'title': '❌ スタッフ数が絶対的に不足しています',
                'message': f'現在: {self.num_staff}名 → 必要: 最低{min_required}名（早番{self.early_count}名 + 中番{self.middle_count}名 + 遅番{self.late_count_min}名）',
                'suggestion': f'✅ スタッフを{shortage}名以上追加してください',
                'priority': 1
            })
        
        # 2. 中番専任が多すぎる
        if nakaban_only_count > self.middle_count:
            excess = nakaban_only_count - self.middle_count
            reasons.append({
                'type': 'critical',
                'title': '❌ 中番専任スタッフが多すぎます',
                'message': f'中番専任: {nakaban_only_count}名 → 中番の必要人数: {self.middle_count}名\n余剰: {excess}名',
                'suggestion': f'✅ 以下のいずれかを実行してください：\n  1. 中番専任を{excess}名減らす（推奨）\n  2. 中番の必要人数を{nakaban_only_count}名に増やす',
                'priority': 1
            })
        
        # 3. 新人が多すぎる（新人同士は同一シフト勤務不可）
        if newbie_count >= 3:
            reasons.append({
                'type': 'warning',
                'title': '⚠️ 新人が多すぎます',
                'message': f'新人: {newbie_count}名 / 全体: {self.num_staff}名\n新人同士は同一シフト勤務できないため、シフトが組みにくくなります。',
                'suggestion': f'✅ 以下のいずれかを実行してください：\n  1. 新人を2名以下に減らす（推奨）\n  2. スタッフ数を{self.num_staff + 2}名以上に増やす',
                'priority': 2
            })
        
        # 4. 希望が多すぎる（個別チェック）
        total_requests = 0
        staff_with_excess_requests = []
        for s in range(self.num_staff):
            requests_off_count = len(self.requests_off.get(s, []))
            preferred_shifts_count = len(self.preferred_shifts.get(s, {}))
            total = requests_off_count + preferred_shifts_count
            total_requests += total
            
            if total > 2:
                staff_name = self.staff_info[s]['name']
                excess = total - 2
                staff_with_excess_requests.append(f'{staff_name}（{total}件 → {excess}件オーバー）')
        
        if staff_with_excess_requests:
            reasons.append({
                'type': 'warning',
                'title': '⚠️ 希望が多すぎるスタッフがいます',
                'message': f'希望休+希望シフトは合計2日までです。\n該当スタッフ: {len(staff_with_excess_requests)}名\n・ ' + '\n・ '.join(staff_with_excess_requests),
                'suggestion': '✅ 各スタッフの希望を2件以内に減らしてください',
                'priority': 2
            })
        
        # 5. 早番・遅番バランス制約の問題（v7で追加）
        if non_nakaban_staff > 0:
            # 早番と遅番のバランスを取るには、十分な勤務日数が必要
            required_working_days_for_balance = 10  # 最低限のバランスを取るために必要な勤務日数
            if working_days < required_working_days_for_balance:
                reasons.append({
                    'type': 'info',
                    'title': 'ℹ️ 早番・遅番バランス制約が厳しい可能性',
                    'message': f'勤務日数: {working_days}日（1ヶ月{self.num_days}日 - 休日{target_off_days}日）\n早番と遅番を均等に振り分けるには、十分な勤務日数が必要です。',
                    'suggestion': f'✅ 休日を{target_off_days - 1}日に減らすことを検討してください',
                    'priority': 3
                })
        
        # 6. 全体的な制約バランスの問題
        avg_working_days_per_staff = (self.num_staff * working_days) / self.num_staff
        min_daily_staff = min_required
        total_shifts_needed = working_days * min_daily_staff
        total_shifts_available = self.num_staff * working_days
        utilization_rate = (total_shifts_needed / total_shifts_available) * 100 if total_shifts_available > 0 else 0
        
        if utilization_rate > 85:
            reasons.append({
                'type': 'warning',
                'title': '⚠️ 制約の余裕が少なすぎます',
                'message': f'シフト充足率: {utilization_rate:.1f}%\n必要シフト数: {total_shifts_needed}\n利用可能シフト数: {total_shifts_available}\n余裕率: {100 - utilization_rate:.1f}%（推奨: 20%以上）',
                'suggestion': f'✅ 以下のいずれかを実行してください：\n  1. スタッフを1-2名追加する（推奨）\n  2. 遅番の最小人数を{self.late_count_min - 1}名に減らす\n  3. 希望休・希望シフトを全体的に減らす',
                'priority': 2
            })
        
        # 7. 一般的な提案（原因が特定できない場合）
        if not reasons:
            reasons.append({
                'type': 'general',
                'title': '⚠️ 制約条件が厳しすぎます',
                'message': f'以下の現状を確認してください：\n・ スタッフ数: {self.num_staff}名\n・ 必要人数: 早番{self.early_count}名、中番{self.middle_count}名、遅番{self.late_count_min}-{self.late_count_max}名\n・ 新人: {newbie_count}名、中番専任: {nakaban_only_count}名\n・ 勤務日数: {working_days}日、休日: {target_off_days}日',
                'suggestion': '✅ 以下を試してください（優先順）：\n  1. スタッフを1-2名追加する\n  2. 希望休・希望シフトを減らす（各スタッフ2件以内）\n  3. 新人を2名以下に減らす\n  4. 中番専任を減らす\n  5. 遅番の最小人数を減らす',
                'priority': 3
            })
        
        # 優先順位でソート
        reasons.sort(key=lambda x: x.get('priority', 999))
        
        return reasons

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_schedule():
    try:
        data = request.json
        
        month = int(data['month'])
        year = int(data['year'])
        num_staff = int(data['num_staff'])
        
        # シフト人数を取得
        early_count = int(data.get('early_count', 2))
        middle_count = int(data.get('middle_count', 1))
        late_count_min = int(data.get('late_count_min', 2))
        late_count_max = int(data.get('late_count_max', 3))
        balance_tolerance = int(data.get('balance_tolerance', 2))  # 早番・遅番バランス許容差
        
        staff_info = {}
        requests_off = {}
        preferred_shifts = {}
        holidays = {}
        
        for i, staff in enumerate(data['staff_list']):
            staff_info[i] = {
                'name': staff['name'],
                'is_newbie': staff.get('is_newbie', False),
                'nakaban_only': staff.get('nakaban_only', False)
            }
            
            # 希望休
            if staff.get('requests_off'):
                requests_off[i] = [int(d) for d in staff['requests_off'].split(',') if d.strip()]
            
            # 希望シフト
            if staff.get('preferred_shifts'):
                preferred_shifts[i] = {}
                for pref in staff['preferred_shifts'].split(','):
                    if ':' in pref:
                        day, shift = pref.split(':')
                        preferred_shifts[i][int(day.strip())] = shift.strip()
            
            # 有給
            if staff.get('holidays'):
                holidays[i] = [int(d) for d in staff['holidays'].split(',') if d.strip()]
        
        scheduler = ShiftScheduler(month, year, num_staff, staff_info, requests_off, preferred_shifts, holidays, early_count, middle_count, late_count_min, late_count_max, balance_tolerance)
        schedule, success, error_reasons, relaxation_info = scheduler.create_schedule_with_relaxation()
        
        if success:
            # スケジュールを整形
            result = []
            for staff_id, shifts in schedule.items():
                result.append({
                    'name': staff_info[staff_id]['name'],
                    'shifts': shifts
                })
            
            response_data = {
                'success': True,
                'schedule': result,
                'num_days': scheduler.num_days
            }
            
            # 制約緩和が適用された場合、その情報を追加
            if relaxation_info:
                response_data['relaxation_info'] = relaxation_info
            
            return jsonify(response_data)
        else:
            return jsonify({'success': False, 'message': 'シフトを作成できませんでした。', 'reasons': error_reasons})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'エラーが発生しました: {str(e)}'})

@app.route('/export', methods=['POST'])
def export_excel():
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from datetime import datetime as dt
        
        data = request.json
        schedule = data['schedule']
        num_days = data['num_days']
        month = int(data['month'])
        year = int(data['year'])
        
        # DataFrameを作成
        df_data = {}
        df_data['スタッフ名'] = [staff['name'] for staff in schedule]
        
        # 曜日付きのカラム名を作成
        days_of_week = ['日', '月', '火', '水', '木', '金', '土']
        column_names = []
        for day in range(1, num_days + 1):
            date = dt(year, month, day)
            day_of_week = days_of_week[date.weekday() if date.weekday() != 6 else 0]
            if date.weekday() == 6:  # 日曜日は0じゃなく6で計算
                day_of_week = '日'
            column_names.append(f'{month}/{day}\n{day_of_week}曜')
            df_data[column_names[-1]] = [staff['shifts'][day - 1] for staff in schedule]
        
        df = pd.DataFrame(df_data)
        
        # Excelファイルとして出力
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='シフト表', index=False)
            
            # ワークシートを取得
            workbook = writer.book
            worksheet = writer.sheets['シフト表']
            
            # 色の定義
            early_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # 早番
            middle_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')  # 中番
            late_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')  # 遅番
            off_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # 休み
            
            sunday_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # 日曜
            saturday_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')  # 土曜
            
            sunday_font = Font(color='CC0000', bold=True)  # 日曜のフォント
            saturday_font = Font(color='0066CC', bold=True)  # 土曜のフォント
            
            # ヘッダーの曜日に色を付ける
            for col_idx, col_name in enumerate(column_names, start=2):  # 2列目から始まる
                cell = worksheet.cell(row=1, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 曜日をチェック
                day_num = int(col_name.split('/')[1].split('\\n')[0])
                date = dt(year, month, day_num)
                
                if date.weekday() == 6:  # 日曜日
                    cell.fill = sunday_fill
                    cell.font = sunday_font
                elif date.weekday() == 5:  # 土曜日
                    cell.fill = saturday_fill
                    cell.font = saturday_font
            
            # シフトのセルに色を付ける
            for row_idx in range(2, len(schedule) + 2):  # 2行目からデータ
                for col_idx in range(2, num_days + 2):  # 2列目からシフトデータ
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    value = str(cell.value)
                    
                    if value.startswith('OFF'):
                        cell.fill = off_fill
                    elif value == '早番':
                        cell.fill = early_fill
                    elif value == '中番':
                        cell.fill = middle_fill
                    elif value == '遅番':
                        cell.fill = late_fill
            
            # 列幅を調整
            worksheet.column_dimensions['A'].width = 15
            for col_idx in range(2, num_days + 2):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 12
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'シフト表_{year}年{month}月.xlsx'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'エラーが発生しました: {str(e)}'})

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
